import hdbscan
from driving.mmd_critic.run_digits_new import run
import openpyxl
import time
import os
import argparse
from numpy import arange
import random
import math
from driving.driving_models import *
import numpy as np
from driving.utils import *


def get_score(x_test, y_test, model):
    pred = model.predict(x_test).reshape(-1)
    true_acc = np.sum(np.square(pred - y_test)) / x_test.shape[0]
    print('Test accuracy:', true_acc)
    return true_acc


def get_ds(countlist,res_get_s, sample_size,X_test,Y_test, X_test2, Y_test2, res,ds, my_model,acc):

    #在非异常点中采样
    len_nonnoise = len(X_test) - countlist[0]
    for key in res:
        b = []
        if len(res[key]) > (len_nonnoise/sample_size):
            #mmd方法采样
            for num in range(int(round(len(res[key]) / (len_nonnoise/sample_size)))):
                b.append(res[key][res_get_s[key][num]])
        else:
            b.append(res[key][res_get_s[key][0]])


        print(len(b))
        for i in range(len(b)):
            X_test2.append(X_test[b[i]])
            Y_test2.append(Y_test[b[i]])

    X_test4 = np.array(X_test2)
    Y_test4 = np.array(Y_test2)

    print(X_test4.shape, Y_test4.shape)
    score = get_score(X_test4, Y_test4, my_model)
    ds.append(np.abs(score - acc))
    # ds.append(np.abs(score - 0.09660315929610656))
    # ds.append(np.abs(score - 0.08176111833886715))


def get_ds_random(countlist,res_get_s, sample_size,X_test,Y_test, X_test2, Y_test2, res,ds, my_model,acc):

    # random.seed(2)

    #在非异常点中采样
    len_nonnoise = len(X_test) - countlist[0]
    for key in res:
        b = []
        if len(res[key]) > (len_nonnoise/sample_size):
            #随机采样
            b = random.sample(res[key], int(round(len(res[key]) / (len_nonnoise/sample_size))))
        else:
            b = random.sample(res[key], 1)

        print(len(b))
        for i in range(len(b)):
            X_test2.append(X_test[b[i]])
            Y_test2.append(Y_test[b[i]])

    X_test4 = np.array(X_test2)
    Y_test4 = np.array(Y_test2)
    print(X_test4.shape, Y_test4.shape)
    score = get_score(X_test4, Y_test4, my_model)
    ds.append(np.abs(score[1] - acc))


def get_std1(X_test, Y_test, a_unoise, countlist,res,label_noise, first_noise,res_get_s,my_model,acc):
    dss = []

    for j in range(30,181,5):
        ds = []

        X_test2 = []
        Y_test2 = []

        len_noise = j*(1-a_unoise)

        print(j)
        #adaptive random
        X_test2.append(X_test[label_noise[first_noise]])
        X_test2.append(X_test[label_noise[np.argmax(dis[first_noise])]])
        Y_test2.append(Y_test[label_noise[first_noise]])
        Y_test2.append(Y_test[label_noise[np.argmax(dis[first_noise])]])

        pre_num = []
        pre_num.append(first_noise)
        pre_num.append(np.argmax(dis[first_noise]))
        while len(X_test2) < len_noise:
            mins = []
            for i in range(len(label_noise)):
                if i not in set(pre_num):
                    min_info = [float('inf'), 0, 0]
                    for l in pre_num:
                        if dis[i][l] < min_info[0]:
                            min_info[0] = dis[i][l]
                            min_info[1] = i
                            min_info[2] = l
                        mins.append(min_info)
            maxnum = 0
            X_test2.append(X_test[label_noise[mins[0][1]]])
            Y_test2.append(Y_test[label_noise[mins[0][1]]])
            pre_num.append(mins[0][1])
            for i in mins:
                if i[0] > maxnum:
                    X_test2[-1] = X_test[label_noise[i[1]]]
                    Y_test2[-1] = Y_test[label_noise[i[1]]]
                    pre_num[-1] = i[1]

        print("异常点挑选个数：", len(X_test2))
        get_ds(countlist, res_get_s, j*a_unoise, X_test, Y_test, X_test2, Y_test2, res, ds, my_model,acc)

        print(ds)
        ds_mean = np.sqrt(np.mean(np.square(ds), axis=0))
        print(ds_mean)
        dss.append(ds_mean)

    print(dss)
    return dss


def get_std1_random(X_test, Y_test, a_unoise, countlist,res,label_noise, first_noise,res_get_s,my_model,acc):
    dss = []

    for j in range(30,181,5):
        ds = []
        # X_test2 = []
        # Y_test2 = []
        len_noise = j*(1-a_unoise)
        # print(j)

        for c in range(50):
            X_test2 = []
            Y_test2 = []

            random_noise = random.sample(label_noise, int(len_noise))
            for i in range(len(random_noise)):
                X_test2.append(X_test[random_noise[i]])
                Y_test2.append(Y_test[random_noise[i]])

            print("异常点挑选个数：", len(X_test2))
            get_ds(countlist, res_get_s, j*a_unoise, X_test, Y_test, X_test2, Y_test2, res, ds, my_model,acc)
        print(ds)
        ds_mean = np.sqrt(np.mean(np.square(ds), axis=0))
        print(ds_mean)
        dss.append(ds_mean)

    print(dss)
    return dss


def load_data():
    path = os.path.join(basedir,'testing/final_example.csv')
    temp = np.loadtxt(path, delimiter=',', dtype=np.str, skiprows=(1))
    names = list(temp[:, 0])
    test = []
    label = []
    for i in range(len(names)):
        n = names[i]
        path = 'testing/center/' + n + '.jpg'
        path = os.path.join(basedir, path)
        test.append(preprocess_image(path))
        label.append(float(temp[i, 1]))
    test = np.array(test)
    test = test.reshape(test.shape[0], 100, 100, 3)
    label = np.array(label)
    return test, label


def add_black(temp, gradients):
    rect_shape = (30, 30)
    for i in range(temp.shape[0]):
        orig = temp[i].reshape(1, 100, 100, 3)
        grad = gradients[i].reshape(1, 100, 100, 3)
        start_point = (
            random.randint(0, grad.shape[1] - rect_shape[0]), random.randint(0, grad.shape[2] - rect_shape[1]))
        new_grads = np.zeros_like(grad)
        patch = grad[:, start_point[0]:start_point[
            0] + rect_shape[0], start_point[1]:start_point[1] + rect_shape[1]]
        new_grads[:, start_point[0]:start_point[0] + rect_shape[0],
                      start_point[1]:start_point[1] + rect_shape[1]] = -np.ones_like(patch)
        orig = orig + 100 * new_grads
        temp[i] = orig.reshape(100, 100, 3)
    return temp

def get_acc(exp_id):
    acc_dict = {"driving_drop":0.08176111833886715,
                'driving_orig':0.09660315929610656}
    return acc_dict[exp_id]

if __name__=="__main__":
    basedir = os.path.abspath(os.path.dirname(__file__))

    """Parser of command args"""
    parse = argparse.ArgumentParser()
    parse.add_argument("--exp_id", type=str, help="exp_identifiers")
    parse.add_argument("--select_layer_idx", type=int, help="selected feature layer")
    parse.add_argument("--dec_dim", type=int, help='decomposition dim')
    parse.add_argument("--min_samples", type=int,help='min samples cluster')
    parse.add_argument("--min_cluster_size", type=int, help='min_cluster_size')

    console_flags, unparsed = parse.parse_known_args(sys.argv[1:])

    select_layer_idx = console_flags.select_layer_idx
    dec_dim = console_flags.dec_dim
    exp_id = console_flags.exp_id
    min_cluster_size = console_flags.min_cluster_size
    min_samples = console_flags.min_samples
    acc = get_acc(exp_id)
    start = time.clock()

    # input image dimensions
    img_rows, img_cols = 100, 100
    input_shape = (img_rows, img_cols, 3)

    # define input tensor as a placeholder
    input_tensor = Input(shape=input_shape)

    # load multiple models sharing same input tensor
    if exp_id == 'driving_drop':
        model = Dave_dropout(input_tensor=input_tensor, load_weights=True)
    elif exp_id == 'driving_orig':
        model = Dave_orig(input_tensor=input_tensor, load_weights=True)
    else:
        raise Exception("no such model {}".format(exp_id))

    print(model.summary())
    # preprocess the data set
    test, label = load_data()
    print("data loaded!")

    X_test = test.copy()
    Y_test = label.copy()
    print(len(X_test))

    dense_layer_model = Model(inputs=model.input, outputs=model.layers[select_layer_idx].output)
    dense_output = dense_layer_model.predict(X_test)
    print(dense_output.shape)
    # from sklearn.preprocessing import MinMaxScaler
    #
    # minMax = MinMaxScaler()
    # dense_output = minMax.fit_transform(dense_output)
    # print(dense_output)
    # dense_output = dense_output.fillna(0)
    from sklearn import preprocessing
    dense_output = preprocessing.normalize(dense_output)

    if exp_id == 'driving_drop':
        from sklearn.decomposition import FastICA
        fica = FastICA(n_components=dec_dim)

    # dense_output = (dense_output - np.mean(dense_output, axis=0)) / np.std(dense_output, axis=0)
    # print(np.isnan(dense_output).any())  # this prints False
    # print(np.isinf(dense_output).any())  # this prints False

        dense_output = fica.fit_transform(dense_output)

    clusterer = hdbscan.HDBSCAN(min_cluster_size=min_cluster_size, min_samples=min_samples)
    r = clusterer.fit(dense_output)
    labels = r.labels_


    print(labels)
    print(np.max(labels))
    print(np.min(labels))

    y_pred_list = labels.tolist()
    countlist = []

    for i in range(np.min(labels), np.max(labels) + 1):
        countlist.append(y_pred_list.count(i))

    print(countlist)
    print(np.sort(countlist))
    print(np.argsort(countlist))

    label_noise = []
    for i, l in enumerate(labels):
        if l == -1:
            label_noise.append(i)

    res = {}
    for i, l in enumerate(labels):
        if l != -1:
            if l not in res:
                res[l] = []
            res[l].append(i)

    print(len(res[0]))

    for key in res:
        X_test3 = []
        Y_test3 = []
        print(len(res[key]))
        for i in range(len(res[key])):
            X_test3.append(X_test[res[key][i]])
            Y_test3.append(Y_test[res[key][i]])
        X_test3 = np.array(X_test3)
        Y_test3 = np.array(Y_test3)
        score = get_score(X_test3, Y_test3, model)

    dis = np.zeros((len(label_noise), len(label_noise)))
    # dis = {}
    # for i in range(len(label_noise)):
    #     # dis[i] = {}
    #     for j in range(len(label_noise)):
    #         if j != i:
    #             dis[i][j] = math.sqrt(np.power(dense_output[label_noise[i]] - dense_output[label_noise[j]], 2).sum())
    #             # dis[i][j] = math.sqrt(np.power(x_test_de[label_noise[i]] - x_test_de[label_noise[j]], 2).sum())

    # noise_score = []
    # for i, l in enumerate(r.outlier_scores_):
    #     if labels[i] == -1:
    #         noise_score.append(l)
    # noise_score = np.array(noise_score)
    #
    # # outlier_sort = np.argsort(-r.outlier_scores_)
    # first_noise = np.argsort(-noise_score)[0]
    # print(noise_score[first_noise])

    first_noise = 0

    res_get_s = {}

    for key in res:
        temp_dense = []
        for l in res[key]:
            temp_dense.append(dense_output[l])
        temp_dense = np.array(temp_dense)
        temp_label = np.full((len(temp_dense)), key)
        mmd_res, _ = run(temp_dense, temp_label, gamma=0.026, m=min(len(temp_dense), 350), k=0, ktype=0, outfig=None,
                    critoutfig=None, testfile=os.path.join(basedir,'data/a.txt'))
        res_get_s[key] = mmd_res

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for a_unoise in arange(0.8, 0.9, 0.1):
        print(a_unoise)
        dss = get_std1_random(X_test=X_test, Y_test=Y_test, a_unoise=a_unoise, countlist=countlist, res=res,
                       label_noise=label_noise, first_noise=first_noise, res_get_s=res_get_s, my_model=model,acc=acc)

        sheet.cell(row=a_unoise * 10 + 1, column=1).value = a_unoise
        for i in range(len(dss)):
            sheet.cell(row=a_unoise * 10 + 1, column=i + 2).value = dss[i]
        elapsed = (time.clock() - start)
        print("Time used: ", elapsed)
        sheet.cell(row=a_unoise * 10 + 1, column=len(dss)+2).value = elapsed

    #workbook.save(os.path.join(basedir, "result", "{}.xlsx".format(exp_id)))
    if console_flags.dec_dim is None:
        workbook.save(os.path.join(basedir, "result", "{}-sn{}-cs{}.xlsx".format(exp_id, min_samples, min_cluster_size)))
    else:
        workbook.save(os.path.join(basedir, "result", "{}-sn{}-cs{}-dim{}.xlsx".format(exp_id, min_samples,
                                                                                       min_cluster_size,
                                                                                       console_flags.dec_dim)))

    elapsed = (time.clock() - start)
    print("Time used: ", elapsed)

